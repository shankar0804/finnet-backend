"""Bulk campaign-entry importer.

Parses either:
  * an .xlsx workbook where each row has a `link` column (optional) plus
    one or more images embedded inside the row's cells, OR
  * a list of pre-parsed dict rows (from a Google Sheet CSV) each with a
    `link` column.

For every row we try to build a campaign entry the same way the single
endpoint does — scrape the link, OCR any image(s), merge, and let
entry_builder_service.build_entry insert. Rows that can't be completed
(missing creator with no rescue data) are returned in a `skipped` array
so the caller surfaces them in the UI.

Runs on a background thread so the HTTP request returns immediately.
"""

from __future__ import annotations

import io
import logging
import threading
import uuid
from typing import Callable, Optional

from services.post_scraper_service import (
    fetch_post_data,
    UnsupportedPlatformError,
    PostNotFoundError,
)
from services.ocr_service import run_post_ocr_pipeline
from services.entry_builder_service import build_entry

logger = logging.getLogger(__name__)

# In-memory job store — same pattern as bulk_import_service. For a single
# Render instance this is fine; if we ever shard, move to Redis / Supabase.
_JOBS: dict = {}
_JOBS_LOCK = threading.Lock()


def _set(job_id: str, **kwargs):
    with _JOBS_LOCK:
        if job_id in _JOBS:
            _JOBS[job_id].update(kwargs)


def get_job(job_id: str) -> Optional[dict]:
    with _JOBS_LOCK:
        job = _JOBS.get(job_id)
        return dict(job) if job else None


def _new_job() -> str:
    job_id = str(uuid.uuid4())
    with _JOBS_LOCK:
        _JOBS[job_id] = {
            'status': 'queued',
            'created': 0,
            'skipped': [],
            'failed': [],
            'total': 0,
            'processed': 0,
            'current': '',
        }
    return job_id


# ═══════════════════════════════════════════════════════════
# Column name normalization
# ═══════════════════════════════════════════════════════════

_LINK_ALIASES = ('link', 'content_link', 'live_link', 'post_link', 'url', 'content')
_USERNAME_ALIASES = ('username', 'creator_username', 'creator', 'handle', 'ig_handle', 'instagram')
_DELIVERABLE_ALIASES = ('deliverable', 'deliverable_type', 'type', 'content_type', 'asset')
_AMOUNT_ALIASES = ('amount', 'commercials', 'fee', 'cost', 'budget', 'price')
_DATE_ALIASES = ('delivery_date', 'date', 'due_date', 'deadline')
_POC_ALIASES = ('poc', 'point_of_contact', 'assigned_to')
_NOTES_ALIASES = ('notes', 'note', 'remarks')


def _first(d: dict, aliases: tuple) -> str:
    for k in aliases:
        v = d.get(k)
        if v not in (None, ''):
            return str(v).strip()
    return ''


def _coerce_amount(raw) -> float:
    if raw in (None, ''):
        return 0.0
    try:
        return float(str(raw).replace(',', '').replace('₹', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _row_to_overrides(raw: dict) -> dict:
    """Normalize a dict with arbitrary casing / column names to form-style overrides."""
    normed = {k.strip().lower().replace(' ', '_'): v for k, v in raw.items() if k}
    return {
        'content_link': _first(normed, _LINK_ALIASES),
        'creator_username': _first(normed, _USERNAME_ALIASES).lstrip('@'),
        'deliverable_type': _first(normed, _DELIVERABLE_ALIASES),
        'amount': _coerce_amount(_first(normed, _AMOUNT_ALIASES)),
        'delivery_date': _first(normed, _DATE_ALIASES) or None,
        'poc': _first(normed, _POC_ALIASES),
        'notes': _first(normed, _NOTES_ALIASES),
    }


# ═══════════════════════════════════════════════════════════
# Excel + embedded image parsing
# ═══════════════════════════════════════════════════════════


def parse_xlsx_with_images(file_bytes: bytes) -> list:
    """Return a list of {row: N, data: {col: val}, images: [bytes,...]}.

    Expects the first sheet, with the first row as headers. Embedded images
    (PNG/JPG pasted into a cell) are picked up via each image's anchor and
    grouped with the row they visually sit on.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    if ws is None:
        return []

    # ─── Header row ───
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1), []):
        headers.append(str(cell.value).strip() if cell.value is not None else '')

    # ─── Data rows ───
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None and str(v).strip() != '' for v in row):
            continue  # skip blank rows
        data = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                data[headers[i]] = val
        rows.append({'row': r_idx, 'data': data, 'images': []})

    # ─── Embedded images ───
    # openpyxl stashes images on the worksheet as ws._images (list of OpenpyxlImage).
    # Each image has .anchor._from.row (0-indexed) — we map it back onto our 1-indexed rows.
    images = getattr(ws, '_images', []) or []
    row_to_index = {row['row']: idx for idx, row in enumerate(rows)}
    for img in images:
        try:
            anchor_row_0 = img.anchor._from.row  # 0-indexed
            anchor_row_1 = anchor_row_0 + 1      # match openpyxl's 1-indexed row numbers
            idx = row_to_index.get(anchor_row_1)
            if idx is None:
                continue
            # Pull raw bytes — the image's ._data is a callable in newer openpyxl.
            data_attr = getattr(img, '_data', None)
            raw = data_attr() if callable(data_attr) else data_attr
            if raw:
                rows[idx]['images'].append(raw)
        except Exception as e:
            logger.warning(f'[BulkEntries] Skipping unparseable embedded image: {e}')

    return rows


# ═══════════════════════════════════════════════════════════
# Per-row processor (shared by xlsx + sheet paths)
# ═══════════════════════════════════════════════════════════


def _process_single_row(
    campaign_id: str,
    row_num: int,
    overrides: dict,
    images: list,
    scrape_acquire: Optional[Callable] = None,
    scrape_release: Optional[Callable] = None,
) -> dict:
    """Scrape + OCR + build_entry for one row. Returns an aggregate entry."""
    link = overrides.get('content_link') or ''
    scraped = None
    ocr = None
    platform_hint = None

    # 1. Scrape (if link present)
    if link:
        acquired = False
        if scrape_acquire is not None:
            try:
                acquired = bool(scrape_acquire())
            except Exception:
                acquired = False
            if not acquired:
                return {
                    'row': row_num,
                    'status': 'failed',
                    'reason': 'Scraper slots busy — timed out waiting for a slot.',
                }
        try:
            scraped = fetch_post_data(link)
            platform_hint = scraped.get('platform')
        except UnsupportedPlatformError as e:
            return {'row': row_num, 'status': 'failed', 'reason': f'Unsupported URL: {e}'}
        except PostNotFoundError as e:
            logger.warning(f'[BulkEntries] Row {row_num}: scrape empty for {link}: {e}')
        except Exception as e:
            logger.warning(f'[BulkEntries] Row {row_num}: scrape error for {link}: {e}')
        finally:
            if acquired and scrape_release is not None:
                try:
                    scrape_release()
                except Exception:
                    pass

    # 2. OCR (fills gaps) — take the first image that parses, then stop.
    for img_bytes in images or []:
        try:
            res = run_post_ocr_pipeline(img_bytes)
            ocr = res.get('result')
            break
        except Exception as e:
            logger.warning(f'[BulkEntries] Row {row_num}: OCR failed on image: {e}')
            continue

    # 3. Build entry
    result = build_entry(
        campaign_id=campaign_id,
        scraped=scraped,
        ocr=ocr,
        overrides=overrides,
        platform_hint=platform_hint,
    )
    result['row'] = row_num
    return result


def process_rows(
    campaign_id: str,
    rows: list,
    job_id: Optional[str] = None,
    scrape_acquire: Optional[Callable] = None,
    scrape_release: Optional[Callable] = None,
    notify: Optional[Callable] = None,
) -> dict:
    """Run the single-row processor for every row. Returns {created, skipped, failed}.

    `rows` is a list of either:
      - {'row': N, 'data': {...}, 'images': [bytes,...]} (from parse_xlsx_with_images)
      - {'row': N, 'data': {...}} (from CSV/Sheet — no images)
    """
    created = 0
    skipped = []
    failed = []
    total = len(rows)
    _set(job_id, status='running', total=total) if job_id else None

    for i, row in enumerate(rows):
        data = row.get('data') or {}
        images = row.get('images') or []
        overrides = _row_to_overrides(data)

        if job_id:
            _set(
                job_id,
                processed=i,
                current=overrides.get('creator_username') or overrides.get('content_link') or f'row {row["row"]}',
            )
        if notify:
            try:
                notify(f'Processing row {i + 1}/{total}')
            except Exception:
                pass

        try:
            result = _process_single_row(
                campaign_id=campaign_id,
                row_num=row.get('row', i + 2),
                overrides=overrides,
                images=images,
                scrape_acquire=scrape_acquire,
                scrape_release=scrape_release,
            )
        except Exception as e:
            logger.exception(f'[BulkEntries] row {row.get("row")} crashed: {e}')
            failed.append({'row': row.get('row'), 'error': str(e)})
            continue

        if result['status'] == 'created':
            created += 1
        elif result['status'] == 'skipped':
            skipped.append({
                'row': result['row'],
                'username': result.get('missing_creator') or overrides.get('creator_username'),
                'reason': result.get('reason'),
                'platform': result.get('platform'),
            })
        else:
            failed.append({'row': result['row'], 'error': result.get('reason')})

    summary = {'created': created, 'skipped': skipped, 'failed': failed, 'total': total}
    if job_id:
        _set(job_id, status='done', processed=total, summary=summary)
    return summary


# ═══════════════════════════════════════════════════════════
# Background launcher — returns immediately with a job_id
# ═══════════════════════════════════════════════════════════


def start_background(
    campaign_id: str,
    rows: list,
    scrape_acquire: Optional[Callable] = None,
    scrape_release: Optional[Callable] = None,
) -> str:
    job_id = _new_job()

    def _run():
        try:
            process_rows(
                campaign_id=campaign_id,
                rows=rows,
                job_id=job_id,
                scrape_acquire=scrape_acquire,
                scrape_release=scrape_release,
            )
        except Exception as e:
            logger.exception(f'[BulkEntries] job {job_id} crashed: {e}')
            _set(job_id, status='failed', error=str(e))

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return job_id
